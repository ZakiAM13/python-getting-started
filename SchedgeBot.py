import openpyxl, datetime, smtplib, imaplib, pyzmail, time, re
from openpyxl.cell import get_column_letter, column_index_from_string
from unidecode import unidecode
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def GetSched(day, n):
    for i in range(1, 12):
        done = False
        col = get_column_letter(column_index_from_string((w.coordinate)[0])+n)
        cls = col+str(int((w.coordinate)[1:])+i)
        time = 'A'+str(int((w.coordinate)[1:])+i)
        for range_ in sheet.merged_cell_ranges:
                merged_cells = list(openpyxl.utils.rows_from_range(range_))
                for row in merged_cells:
                    if cls in row:
                        if cls == merged_cells[0][0] and sheet[cls].value != None and "Week" not in sheet[cls].value:
                            apd = [list(filter(None, sheet[cls].value.split('\n'))), sheet[time].value]
                            for i in range(1, len(merged_cells)):
                                apd.append(sheet[time[0]+(merged_cells[i][0])[1:]].value)
                            day.append(apd)
                        done = True
        if sheet[cls].value != None and "Week" not in sheet[cls].value and done == False:
            day.append([list(filter(None, sheet[cls].value.split('\n'))), sheet[time].value])
    for i in day:
        if len(i)>2:
            day[day.index(i)] = [i[0], ((i[1].split("-"))[0]+'-'+(i[-1].split("-"))[1])]
def printer(day):
    sch = '<p><strong>'
    for i in day:
        for t in i:
            if type(t) is list:
                for z in t:
                    sch += z + ' ' 
            else:
                sch += "</strong>from " + t.replace("-", " to ") +"<br/><strong>"
    return sch+"</p></strong>"

# Configure the program by setting some variables.
MY_EMAIL = os.environ['EMAIL_LIST']
BOT_EMAIL = os.environ['BOT_MAIL']
BOT_EMAIL_PASSWORD = os.environ['BOT_PASS']
IMAP_SERVER = 'imap.gmail.com'
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465

wb = openpyxl.load_workbook('Med I schedule Jan-Jun.xlsx')
sheet = wb.get_sheet_by_name('schedule')

dates = ['3','16','30','44','58','72','85', '99', '111']
datecells = []
for i in dates:
    datecells.append(tuple(sheet['B'+i:'S'+i])[0])

week = (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y-%m-%d")

for i in datecells:
	for d in range(0, len(i)):
		if i[d].value.strftime("%Y-%m-%d") == week:
			w = i[d]
			break
Mon = []
Tue = []
Wed = []
Thur = []
Fri = []
Sat = []

GetSched(Mon, 0)
GetSched(Tue, 1)
GetSched(Wed, 2)
GetSched(Thur, 3)
GetSched(Fri, 4)
GetSched(Sat, 5)

# Send an email response about the task.
#responseBody = "Monday:\n"+printer(Mon)+"Tuesday:\n"+printer(Tue)+"Wednesday:\n"+printer(Wed)+"Thursday:\n"+printer(Thur)+"Friday:\n"+printer(Fri)+"Saturday:\n"+printer(Sat) 
#body = re.sub(r'[^\x00-\x7F]+',' ', responseBody)
body="""\
   <html>
     <head></head>
     <body>
        <h2><span style="text-decoration: underline; color: #ff0000;"><strong>Monday:</strong></span></h2>
        {}
        <h2><span style="text-decoration: underline; color: #ff0000;"><strong>Tuesday:</strong></span></h2>
        {}
        <h2><span style="text-decoration: underline; color: #ff0000;"><strong>Wednesday:</strong></span></h2>
        {}
        <h2><span style="text-decoration: underline; color: #ff0000;"><strong>Thursday:</strong></span></h2>
        {}
        <h2><span style="text-decoration: underline; color: #ff0000;"><strong>Friday:</strong></span></h2>
        {}
        <h2><span style="text-decoration: underline; color: #ff0000;"><strong>Saturday:</strong></span></h2>
        {}
     </body>
   </html>
   """.format(printer(Mon),printer(Tue),printer(Wed),printer(Thur),printer(Fri),printer(Sat))
msg = MIMEMultipart('alternative')
msg['To'] = ', '.join(MY_EMAIL)
msg['From'] = "Schedule Bot <mp.polo13@gmail.com>"
msg['Subject'] = 'Schedule for ' + sheet[(w.coordinate)[0]+str(int((w.coordinate)[1:])-2)].value
msg.attach(MIMEText(body, 'html'))
smtpCli = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
smtpCli.ehlo()
smtpCli.login(BOT_EMAIL, BOT_EMAIL_PASSWORD)
smtpCli.sendmail(BOT_EMAIL, MY_EMAIL, msg.as_string())
smtpCli.quit()


                
                    
    
                
